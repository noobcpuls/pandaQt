import os
import string
from turtle import title
from typing import overload
from PyQt5.QtGui import QIcon, QStandardItemModel
from PyQt5.uic.uiparser import QtCore, QtWidgets
import numpy as np
import pandas as pd
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic
from openpyxl import load_workbook
import copy
import matplotlib as mpl
import matplotlib.pyplot as plt
import platform

if platform.system() == 'Darwin': #맥
        plt.rc('font', family='AppleGothic') 
elif platform.system() == 'Windows': #윈도우
        plt.rc('font', family='Malgun Gothic') 

mpl.rcParams['axes.unicode_minus'] = False

def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

form = resource_path('panda_qt.ui')
form_class = uic.loadUiType(form)[0]

def render_mpl_table(data, title, col_width=3.0, row_height=0.625, font_size=11,
                     header_color='#717ED1', row_colors=['#FFFFFF', 'w'], edge_color='#000000',
                     bbox=[0, 0, 1, 1], header_columns=0, 
                     ax=None, **kwargs):
    if ax is None:
        fig, ax = plt.subplots(figsize=(8.27, 11.69))
        ax.axis('off')
        ax.set_title(title ,pad=20)
    mpl_table = ax.table(cellText=data.values, cellLoc='center', bbox=bbox, colLabels=data.columns, **kwargs)
    mpl_table.auto_set_font_size(False)
    mpl_table.set_fontsize(font_size)

    for k, cell in mpl_table._cells.items():
        cell.set_edgecolor(edge_color)
        if k[0] == 0 or k[1] < header_columns:
            cell.set_text_props(weight='bold', color='w')
            cell.set_facecolor(header_color)
        else:
            cell.set_facecolor(row_colors[k[0]%len(row_colors) ])
    return ax.get_figure(), ax

def isWindow():
    if platform.system() == 'Darwin':
        return False
    elif platform.system() == 'Windows':
        return True
    else:
        return False

class Data(): 
    def __init__(self):
        self.__students = []
        self.__maxCor = 0
        d = {"이름": self.__students, "맞은 갯수": [], "점수": [], "순위": []}
        self.__df = pd.DataFrame(data=d)

    #getter
    def getStudents(self):
        return self.__students

    def getMaxCor(self):
        return self.__maxCor

    def getDf(self):
        return self.__df

    #setter
    def setStudnets(self, value):
        self.__students.append(value)
        data_to_insert = {"이름": value, "맞은 갯수": -1, "점수": 0, "순위": 0}
        self.__df = self.__df.append(data_to_insert, ignore_index=True)

    def setMaxCor(self, value):
        self.__maxCor = value

    #calc function
    def calcScore(self):
        for i in range(len(self.__df.index)):
            correct = self.__df.iloc[i, 1]
            if correct != -1 and correct != '미응시':
                score = int(correct) / self.__maxCor
                score2 = score * 100
                self.__df.iloc[i, 2] = round(score2, 1)
            else:
                self.__df.iloc[i, 1] = '미응시'
                self.__df.iloc[i, 3] = 0
        
        self.__df.sort_values(by=['점수'], axis=0, ascending=False, inplace=True)

        grade = 1
        for i in range(len(self.__df.index)):
            if self.__df.iloc[i, 1] != '미응시':
                self.__df.iloc[i, 3] = round(grade, 0)
            else:
                self.__df.iloc[i, 3] = 0
            grade += 1

    def calcAvg(self):
        scoreList = self.__df.점수
        total = 0
        stu = 0

        for i in range(len(scoreList)):
            score = self.__df.iloc[i, 2] if self.__df.iloc[i, 1] != '미응시' else -1
            if score != -1:
                total += scoreList[i]
                stu += 1
        
        avg = total / stu

        return avg

    def sortStudent(self):
        self.__df.sort_values(by=['이름'], axis=0, ascending=True, inplace=True)

    #delete function
    def delStudent(self, value):
        self.__df.drop(index=self.__df.loc[self.__df.이름 == value].index, inplace=True)

    def saveToExcel(self, grade, value):

        path = f'{grade}.xlsx'

        if not os.path.isfile(path):
            self.__df.to_excel(f'{grade}.xlsx', sheet_name=f'{value}', index=False)
            return

        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book

        self.__df.to_excel(writer, sheet_name=f'{value}', index=False)
        writer.save()
        writer.close()

    def readExcel(self, grade, name):
        x = pd.read_excel(f'{grade}.xlsx', sheet_name=f'{name}')
        self.__df = x

    def hideName(self):
        scoreList = self.__df.점수.tolist()
        print(scoreList)

        for i in range(len(scoreList)):
            print(scoreList[i])
            if float(scoreList[i]) < 80.0:
                name = self.__df.iloc[i, 0]
                print(name)
                newName = name[0] + '**'
                self.__df.iloc[i, 0] = newName

data = Data()

class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)

        self.setWindowTitle('판다 성적 입력 프로그램')
        icon_path = resource_path('panda_bear_icon_153300.svg')
        self.setWindowIcon(QIcon(icon_path))
        self.show()

        #quit button
        self.quitButton.clicked.connect(QCoreApplication.instance().quit)
        
        #add, modify and set student button
        self.setStudent.clicked.connect(self.addTableItemDialog)
        self.modifyStudent.clicked.connect(self.modifyTableItemDialog)
        self.delStudent.clicked.connect(self.delTableItem)

        #combo box widget
        select = self.comboBox
        select.currentIndexChanged.connect(self.comboBoxChangedSignal)

        #dateEdit widget
        date = self.dateEdit
        date.setDate(QDate.currentDate())

        #list widget
        listW = self.listWidget
        val = select.currentText()
        if os.path.isfile(f'{val}.xlsx'):
            xl = pd.ExcelFile(f'{val}.xlsx')
            for j in range(len(xl.sheet_names)):
                listW.addItem(xl.sheet_names[j])

        listW.itemDoubleClicked.connect(self.readItemToList)

        #CRUD Item to List button
        addBtn = self.addItemBtn
        addBtn.clicked.connect(self.addItemToList)
        delBtn = self.delItemBtn
        delBtn.clicked.connect(self.delItemToList)
        # updBtn = self.overWriteBtn
        # updBtn.clicked.connect(self.updateItemToList)
        loadBtn = self.loadFileBtn
        loadBtn.clicked.connect(self.loadFile)

        #correct number lineEdit
        MSinput = self.maxScoreInput
        MSinput.textChanged.connect(self.changedScoreSignal)

        #sort button
        sortStuBtn = self.sortByStudentName
        sortScoBtn = self.sortByScore
        sortScoBtn.clicked.connect(self.sortScoreSignal)
        sortStuBtn.clicked.connect(self.sortStudentSignal)
        sortGraBtn = self.sortByGrade
        sortGraBtn.clicked.connect(self.sortGradeSignal)
        self.printSheet.clicked.connect(self.printSheetSignal)

        #table widget
        tableHeader = self.tableWidget.horizontalHeader()
        tableHeader.setSectionResizeMode(0, QHeaderView.Stretch)
        tableHeader.setSectionResizeMode(1, QHeaderView.Stretch)
        tableHeader.setSectionResizeMode(2, QHeaderView.Stretch)
        table = self.tableWidget
        table.cellChanged.connect(self.changedTableSignal)

        #clear button
        clear = self.clearBtn
        clear.clicked.connect(self.clearTable)

    def addTableItemDialog(self):
        text, ok = QInputDialog.getText(self, '입력할 학생 정보 입력', '이름: ')

        if ok:
            table = self.tableWidget
            row_count = table.rowCount()
            table.insertRow(row_count)
            table.setVerticalHeaderItem(row_count, QTableWidgetItem(str(text)))
            data.setStudnets(str(text))
    
    def modifyTableItemDialog(self):
        table = self.tableWidget
        cur_row = table.currentRow()
        item = table.verticalHeaderItem(cur_row)

        if type(item) != QTableWidgetItem:
            QMessageBox.about(self, '경고', '선택된 학생이 없습니다.')
            return

        text, ok = QInputDialog.getText(self, '변경할 학생 정보 입력', '이름: ')

        if ok:
            data.getDf().iloc[cur_row, 0] = str(text)

            table.setVerticalHeaderItem(cur_row, QTableWidgetItem(str(text)))

    def delTableItem(self):
        table = self.tableWidget
        cur_row = table.currentRow()
        value = table.verticalHeaderItem(cur_row)
        
        if type(value) == QTableWidgetItem:
            data.delStudent(value.text())
            print(value.text())

        table.removeRow(cur_row)

    def alertDialog(self, alertText):
        QMessageBox.about(self, "경고", alertText)

    def addItemToList(self):
        text = self.titleEdit.text()
        pandaList = self.listWidget
        pandaDate = self.dateEdit.date().toString(Qt.ISODate)
        pandaGrade = self.comboBox.currentText()

        if text == '':
            itemText = f'{pandaGrade} {pandaDate} 성적표'
            pandaList.addItem(itemText)
            data.saveToExcel(pandaGrade, itemText)
        else:
            pandaList.addItem(text)
            data.saveToExcel(pandaGrade, text)

    def readItemToList(self):
        pandaList = self.listWidget
        pandaGrade = self.comboBox.currentText()
        item = pandaList.currentItem().text()
        file = f'{pandaGrade}.xlsx'

        if os.path.isfile(file):
            try:
                data.readExcel(pandaGrade, item)
            except ValueError:
                QMessageBox.about(self, '경고', '파일을 찾을 수 없습니다.')
            table = self.tableWidget
            r_list = data.getDf().이름
            c_list = ['맞은 갯수', '점수', '순위']

            table.clear()
            table.setRowCount(len(r_list))
            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(c_list)
            table.setVerticalHeaderLabels(r_list)

            for i in range(len(c_list)):
                for j in range(len(r_list)):
                    table.setItem(j, i, QTableWidgetItem(str(data.getDf().iloc[j, i+1])))

        else:
            QMessageBox.about(self, '경고', '파일을 찾을 수 없습니다.')

    def delItemToList(self):
        pandaList = self.listWidget
        cur_row = pandaList.currentRow()
        text = pandaList.currentItem().text()
        pandaGrade = self.comboBox.currentText()
        file = f'{pandaGrade}.xlsx'

        if os.path.isfile(file):
            book = load_workbook(file)
            if len(book.sheetnames) <= 1:
                QMessageBox.about(self, '경고', '하나 이상의 시트가 있어야 합니다.')
                return
            if text in book.sheetnames:
                book.remove(book[text])
                book.save(file)
        else:
            QMessageBox.about(self, '경고', '파일을 찾을 수 없습니다.')
        pandaList.takeItem(cur_row)
        
    def loadFile(self):
        pandaList = self.listWidget
        file = QFileDialog.getOpenFileName(self, '파일 선택', './', filter='*.xlsx')
        
        if file[0]:
            pandaList.clear()
            xl = pd.ExcelFile(file[0])
            for i in range(len(xl.sheet_names)):
                pandaList.addItem(xl.sheet_names[i])
        else:
            return

    def comboBoxChangedSignal(self):
        pandaList = self.listWidget
        pandaGrade = self.comboBox.currentText()
        file = f'{pandaGrade}.xlsx'

        if os.path.isfile(file):
            pandaList.clear()
            xl = pd.ExcelFile(file)
            for i in range(len(xl.sheet_names)):
                pandaList.addItem(xl.sheet_names[i])
        
        else:
            pandaList.clear()
            return


    def sortStudentSignal(self):
        data.sortStudent()
        table = self.tableWidget
        r_list = data.getDf().이름
        c_list = ['맞은 갯수', '점수', '순위']

        table.clear()
        table.setHorizontalHeaderLabels(c_list)
        table.setVerticalHeaderLabels(r_list)

        for i in range(len(c_list)):
            for j in range(len(r_list)):
                table.setItem(j, i, QTableWidgetItem(str(data.getDf().iloc[j, i+1])))

    def sortScoreSignal(self):
        if data.getMaxCor() <= 0:
            QMessageBox.about(self, '경고', '점수는 0 이상이어야 합니다.')
            return
        
        data.calcScore()
        table = self.tableWidget
        
        r_list = data.getDf().이름
        c_list = ['맞은 갯수', '점수', '순위']
        
        table.clear()
        table.setHorizontalHeaderLabels(c_list)
        table.setVerticalHeaderLabels(r_list)

        for i in range(len(c_list)):
            for j in range(len(r_list)):
                table.setItem(j, i, QTableWidgetItem(str(data.getDf().iloc[j, i+1])))

    def sortGradeSignal(self):
        data.hideName()
        table = self.tableWidget
        
        r_list = data.getDf().이름
        c_list = ['맞은 갯수', '점수', '순위']

        table.clear()
        table.setHorizontalHeaderLabels(c_list)
        table.setVerticalHeaderLabels(r_list)

        for i in range(len(c_list)):
            for j in range(len(r_list)):
                table.setItem(j, i, QTableWidgetItem(str(data.getDf().iloc[j, i+1])))


    def changedScoreSignal(self):
        value = self.maxScoreInput.text()
        if value != '':
            try:
                value2 = int(value)
                data.setMaxCor(value2)
                print(data.getMaxCor())
            except ValueError:
                QMessageBox.about(self, '경고', '숫자만 넣어주세요.')
                return
        else:
            return
        
        
    def changedTableSignal(self):
        value = self.tableWidget.currentItem()
        cur_row = self.tableWidget.currentRow()
        cur_col = self.tableWidget.currentColumn()

        if type(value) == QTableWidgetItem:
            try:
                data.getDf().iloc[cur_row, cur_col + 1] = int(value.text())
            except ValueError:
                QMessageBox.about(self, '경고', '숫자만 넣어주세요.')
                return


    def clearTable(self):
        self.tableWidget.clear()
        self.tableWidget.setRowCount(0)
        col = data.getDf().columns

        self.tableWidget.setHorizontalHeaderLabels(col)

    def printSheetSignal(self):
        df = copy.deepcopy(data.getDf())
        mis_list = []

        if len(df.index) < 1:
            QMessageBox.about(self, '경고', '테이블이 없습니다.')
            return

        if '맞은 갯수' in df.columns:
            df.drop('맞은 갯수', axis=1, inplace=True)
            col_change = ['순위', '이름', '점수']
            df = df.reindex(columns=col_change)

        if len(df.index) <= 30:
            leng = len(df.index)
            missing = 0
            for i in range(leng):
                if df.iloc[i, 0] == 0:
                    missing += 1
                    mis_list.append(df.iloc[i, 1])
                    df = df.drop(df.index[i])
                    print('missing found?!')
            col_len = 30 - leng + missing
            for i in range(col_len):
                df = df.append({'순위': '', '이름': '', '점수': ''}, ignore_index=True)
        else:
            return
        print(mis_list)
        print(df)
        
        d = {"순위": [], "이름": [], "점수": []}
        dfm = pd.DataFrame(data=d)
        for i in range(len(mis_list)):
            dfm = dfm.append({'순위': '미응시', '이름': mis_list[i], '점수': ''}, ignore_index=True)
        dfm.sort_values(by=['이름'], axis=0, ascending=True, inplace=True)
        print(dfm)
        
        df = pd.concat([df, dfm], axis=0, ignore_index=True)
        print(df)

        if df.iloc[-1, 0] != '평균':
            avg = data.calcAvg()
            avgStr = f'{avg:.2f}'
            df = df.append({'순위': '평균', '이름': '', '점수': avgStr}, ignore_index=True)
        
        if self.titleEdit.text() != '':
            title = self.titleEdit.text()
        else:
            title = self.listWidget.currentItem().text()

        fig,ax = render_mpl_table(df, title, header_columns=0, col_width=2.0)
        fig.savefig(f'{title}.pdf')
        
                

if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #WindowClass의 인스턴스 생성
    myWindow = WindowClass() 

    #프로그램 화면을 보여주는 코드
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()